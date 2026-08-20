"""xlsx comparison writer. Phase 4 (Output Agent).

Three sheets (CLAUDE.md §8 Phase 4 + Appendix A.3):
  - Summary : one row per product
  - Variants: one row per SKU across all products (one column per option group)
  - Reviews : one row per review

Freeze the header row, auto-width columns, CNY number format. Never crashes on
zero-review / single-variant products.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config import load_config
from src.models import Product

_CNY = '"¥"#,##0.00'


def _header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _auto_width(ws: Worksheet) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 8), 60)


def _cny(ws: Worksheet, col_idx: int) -> None:
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for c in row:
            c.number_format = _CNY


# Formula-injection prefixes a spreadsheet app parses as a formula (CSV/SSE vector):
# leading '=', '+', '-', '@' — INCLUDING a leading-whitespace-prefixed variant, since
# Excel/Sheets trim whitespace before parsing ('  =1+1' is just as risky as '=1+1').
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _is_formula_risk(value) -> bool:
    """Pure: a string a spreadsheet app would parse as a formula (= + - @), even when
    preceded by leading whitespace. Conservative."""
    return isinstance(value, str) and value.lstrip().startswith(_FORMULA_PREFIXES)


def _neutralize_formulas(ws: Worksheet) -> None:
    """Force every cell a spreadsheet app would parse as a formula to literal TEXT.

    openpyxl auto-turns any string starting with '=' into an executable formula
    (data_type 'f'); product titles/shops/variant labels/review text are Taobao-controlled,
    so a value like '=HYPERLINK(...)' would execute when the buyer opens the workbook.
    We also conservatively force leading '= + - @' (and a whitespace-prefixed variant) to
    TEXT. These writers contain NO intentional formulas, so a blanket neutralization is safe.
    """
    for row in ws.iter_rows():
        for c in row:
            if _is_formula_risk(c.value):
                c.data_type = "s"


def _variant_review_count(product: Product, variant) -> int:
    """Reviews tagged to this variant.

    reviews_by_variant is keyed by the review's full 已购 label. Match a key to this
    variant when ALL of the variant's property values appear in the key (multi-group
    "黑 L"), OR the key is a substring of the variant's joined values (Taobao's
    abbreviated labels). Ambiguous abbreviations attach to each plausible variant
    instead of orphaning to 0 (fixes the multi-group / abbreviated-label bug).
    """
    rbv = product.reviews_by_variant
    values = list(variant.properties.values())
    if not rbv or not values:
        return 0
    # EXACT match only — a single property value, or the full space-joined label.
    # (Substring matching cross-contaminated e.g. "黑" onto "黑色升级版".)
    candidates = set(values) | {" ".join(values)}
    return sum(len(revs) for key, revs in rbv.items() if key in candidates)


def write_xlsx(products: list[Product], filename: str, out_dir: str | None = None) -> str:
    """Write the 3-sheet comparison workbook; return the absolute file path."""
    out_dir = out_dir or load_config().output.dir
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    filename = Path(filename).name or "comparison"   # containment: ignore any path/.. in filename
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    path = Path(out_dir) / filename

    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    _header(ws, ["Title", "Shop", "Min ¥", "Max ¥", "#Variants", "#Reviews", "%Reviews w/ images", "URL"])
    for p in products:
        nrev = len(p.reviews)
        img_pct = round(100 * sum(1 for r in p.reviews if r.has_images) / nrev, 1) if nrev else 0
        mn = p.price_range[0] if p.price_range else None
        mx = p.price_range[1] if p.price_range else None
        ws.append([p.title, p.shop_name, mn, mx, len(p.variants), nrev, img_pct, p.url])
    _cny(ws, 3)
    _cny(ws, 4)

    # --- Variants (one column per discovered option group) ---
    wsv = wb.create_sheet("Variants")
    groups: list[str] = []
    for p in products:
        for v in p.variants:
            for k in v.properties:
                if k not in groups:
                    groups.append(k)
    _header(wsv, ["Product ID", "Title", *groups, "Price ¥", "Stock", "Available", "#Reviews(variant)"])
    for p in products:
        for v in p.variants:
            wsv.append([
                p.product_id, p.title,
                *[v.properties.get(g, "") for g in groups],
                v.price, v.stock, "Yes" if v.available else "No",
                _variant_review_count(p, v),
            ])
    _cny(wsv, 2 + len(groups) + 1)

    # --- Reviews ---
    wsr = wb.create_sheet("Reviews")
    _header(wsr, ["Product ID", "SKU bought", "Rating", "Has images", "Date", "Text (raw 中文)"])
    for p in products:
        for r in p.reviews:
            wsr.append([p.product_id, r.sku_bought, r.rating, "Yes" if r.has_images else "No", r.date, r.text])

    for sheet in (ws, wsv, wsr):
        _auto_width(sheet)
        _neutralize_formulas(sheet)   # formula-injection guard: '='-leading text stays text

    wb.save(path)
    return str(path.resolve())


def write_compare_xlsx(rows: list[dict], filename: str, out_dir: str | None = None) -> str:
    """Write a one-sheet comparison workbook from compare rows (no Product needed).

    Each row is a dict from src.extract.compare._summarize (title/shop/price_range/
    variant_count/price_sample/review_count/subsidy_caveat/url/product_id or an error dict).
    Returns the absolute file path.
    """
    out_dir = out_dir or load_config().output.dir
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    filename = Path(filename).name or "compare"  # containment: ignore any path/.. in filename
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    path = Path(out_dir) / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Compare"
    _header(ws, ["Product ID", "Title", "Shop", "Min ¥", "Max ¥", "#Variants",
                 "Price sample (¥)", "#Reviews", "Subsidy caveat", "URL"])
    for r in rows:
        if "error" in r:
            ws.append([r.get("product_id"), f"ERROR: {r.get('error')}", "", "", "", "", "", "", "", ""])
            continue
        pr = r.get("price_range") or ()
        ws.append([
            r.get("product_id"), r.get("title"), r.get("shop"),
            pr[0] if len(pr) > 0 else None, pr[1] if len(pr) > 1 else None,
            r.get("variant_count"),
            ", ".join(str(p) for p in (r.get("price_sample") or [])),
            r.get("review_count"), r.get("subsidy_caveat"), r.get("url"),
        ])
    _cny(ws, 4)
    _cny(ws, 5)
    _auto_width(ws)
    _neutralize_formulas(ws)   # formula-injection guard for untrusted title/shop/caveat text
    wb.save(path)
    return str(path.resolve())
