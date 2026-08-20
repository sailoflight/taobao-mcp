"""Unit tests for src/inventory.py — dinamic parsing, landed-cost allocation, dedup, categorize."""

import json
import os

from src.inventory import (
    _AllowedRedirectHandler,
    _UrlPolicyError,
    _formula_image_url,
    _formula_link_url,
    _image_url_policy,
    _qty,
    _resolve_out_path,
    _to_num,
    accumulate_dinamic,
    build_xlsx,
    categorize,
    inventory_rows,
)


def _order(oid, day, seller, items, shipping=None, paid=None, status="交易成功"):
    o = {"createDay": day, "seller": seller, "status": status, "items": items}
    if shipping is not None:
        o["shipping"] = shipping
    if paid is not None:
        o["order_paid"] = paid
    return {oid: o}


def _item(title, price, qty=1, variant="", item_id="111", pic="//img.alicdn.com/x.jpg"):
    return {"title": title, "variant": variant, "price": price, "qty": qty,
            "itemId": item_id, "pic": pic, "itemUrl": "https://item.taobao.com/item.htm?id=" + item_id}


# ── numeric coercion ─────────────────────────────────────────────────────────
def test_to_num_and_qty_edge_cases():
    assert _to_num("￥1,371.25") == 1371.25
    assert _to_num("¥6.00") == 6.0
    assert _to_num(None) == 0.0
    assert _to_num("") == 0.0
    assert _qty({"qty": None}) == 1
    assert _qty({"qty": "3"}) == 3
    assert _qty({"qty": 0}) == 1     # never divide by zero / negative
    assert _qty({"qty": -2}) == 1


# ── landed cost ──────────────────────────────────────────────────────────────
def test_landed_single_item_order_reconciles_to_paid():
    # the real-world case that prompted the column: ¥7 product + ¥6 shipping = ¥13 实付款
    by = _order("1001", "2026-06-24", "深圳市鑫达电子",
                [_item("932SQ420DGLF 贴片", "￥7.00")], shipping="￥6.00", paid="￥13.00")
    (r,) = inventory_rows(by, since="2025-01-01")
    assert r["unit"] == 7.0
    assert r["ship"] == 6.0
    assert r["landed_unit"] == 13.0
    assert r["landed_line"] == 13.0


def test_landed_multi_item_allocation_sums_to_paid():
    # shipping is ONE order-level fee spread by quantity across all units
    by = _order("1002", "2026-06-23", "tb44759054",
                [_item("主板", "￥1.50", qty=5, item_id="a"),
                 _item("螺旋桨", "￥2.00", qty=20, item_id="b")],
                shipping="￥7.00", paid="￥54.50")
    rows = inventory_rows(by, since="2025-01-01")
    assert len(rows) == 2
    total_landed = round(sum(r["landed_line"] for r in rows), 2)
    # 1.5*5 + 2*20 = 47.50 product + 7.00 shipping = 54.50 = 实付款
    assert abs(total_landed - 54.50) <= 0.05    # allocation rounding tolerance
    for r in rows:  # per-unit share identical across the order: 7/25 = 0.28
        assert r["landed_unit"] == round(r["unit"] + 0.28, 2)


def test_free_shipping_landed_equals_product():
    by = _order("1003", "2026-06-25", "veromoda官方奥莱旗舰店",
                [_item("上衣", "￥178.03", variant="深棕;L")])   # no shipping field (包邮)
    (r,) = inventory_rows(by, since="2025-01-01")
    assert r["ship"] == 0.0
    assert r["landed_line"] == r["line_total"] == 178.03


# ── filtering / dedup / flags ────────────────────────────────────────────────
def test_since_filter_and_dedup():
    by = {}
    by.update(_order("2001", "2024-12-30", "old", [_item("旧货", "￥1.00")]))
    by.update(_order("2002", "2025-01-02", "new", [_item("新货", "￥2.00"),
                                                   _item("新货", "￥2.00")]))  # duplicate line
    rows = inventory_rows(by, since="2025-01-01")
    assert [r["order_no"] for r in rows] == ["2002"]   # 2024 order excluded, dupe collapsed
    assert len(rows) == 1


def test_custom_link_flag_and_food_kind():
    by = {}
    by.update(_order("3001", "2026-06-22", "某厂家直批",
                     [_item("1元补差价", "￥0.98", qty=2840)]))
    by.update(_order("3002", "2026-06-27", "麦当劳麦乐送(某店)",
                     [_item("商家配送", "￥6.00")]))
    rows = {r["order_no"]: r for r in inventory_rows(by, since="2025-01-01")}
    assert rows["3001"]["custom"] is True              # opaque payment-link line flagged
    assert rows["3002"]["kind"] == "food/local"        # instant delivery marked, not goods
    assert rows["3001"]["kind"] == "goods"


# ── dinamic body parsing ─────────────────────────────────────────────────────
def test_accumulate_dinamic_joins_nodes_by_order_id():
    body = json.dumps({"data": {"data": {
        "shopInfo_42": {"fields": {"orderId": "42", "createDay": "2026-06-24",
                                   "sellerName": "深圳市鑫达电子"}},
        "orderStatus_42": {"fields": {"subTitle": "卖家已发货"}},
        "orderPayment_42": {"fields": {"actualFee": {"value": "￥13.00"},
                                       "pcPostFee": {"value": "￥6.00"}}},
        "orderItemInfo_42_42": {"fields": {"item": {
            "title": "932SQ420DGLF", "skuText": "SM-8", "quantity": 1, "itemId": "874",
            "priceInfo": {"actualTotalFee": "￥7.00"}, "pic": "//img.alicdn.com/p.jpg",
            "itemUrl": "https://item.taobao.com/item.htm?id=874&mi_id=zzz"}}},
    }}})
    by = {}
    accumulate_dinamic(body, by)
    o = by["42"]
    assert o["seller"] == "深圳市鑫达电子" and o["createDay"] == "2026-06-24"
    assert o["shipping"] == "￥6.00" and o["order_paid"] == "￥13.00"
    (it,) = o["items"]
    assert it["variant"] == "SM-8" and it["price"] == "￥7.00"
    assert "mi_id" not in it["itemUrl"]                # tracking param stripped
    accumulate_dinamic(body, by)                       # replay same page → no duplicate items
    assert len(by["42"]["items"]) == 1


# ── categorization fallback ──────────────────────────────────────────────────
def test_categorize_keyword_fallback(tmp_path):
    rows = [
        {"title": "ESP32-C6 SuperMini开发板", "item_id": "1"},
        {"title": "34V 3KW 纯正弦逆变器", "item_id": "2"},
        {"title": "神秘商品XYZ", "item_id": "3"},
    ]
    categorize(rows, prior_xlsx=str(tmp_path / "missing.xlsx"))   # no prior file → keywords
    assert rows[0]["category"] == "MCU / Dev board"
    assert rows[1]["category"] == "Inverter / UPS / Battery"
    assert rows[2]["category"] == "Other"


# ── filename containment (never escape the configured output dir) ─────────────
def test_resolve_out_path_containment(tmp_path):
    out = str(tmp_path / "out")
    cases = {
        "plain.xlsx": "plain.xlsx",
        "/tmp/evil.xlsx": "evil.xlsx",                 # absolute path → basename only
        "../evil.xlsx": "evil.xlsx",                   # traversal → basename only
        "sub/dir/nested.xlsx": "nested.xlsx",          # subdir → basename only
        "": "inventory.xlsx",                          # empty → default name
        None: "inventory.xlsx",
    }
    for given, expect in cases.items():
        got = _resolve_out_path(given, out_dir=out)
        assert got == f"{out}/{expect}", f"{given!r} -> {got}"
        assert os.path.dirname(got) == out              # always inside out_dir
        assert os.path.basename(got) == expect          # only the basename is used


# ── spreadsheet formula injection: neutralize untrusted, keep validated formulas ──
def _inv_row(**over):
    r = {
        "date": "2025-01-02", "seller": "店铺A", "title": "收纳箱", "variant": "特大号",
        "qty": 1, "unit": 10.0, "line_total": 10.0, "ship": 0.0,
        "landed_unit": 10.0, "landed_line": 10.0, "status": "已签收", "order_no": "123",
        "custom": False, "kind": "goods", "pic": "//img.alicdn.com/p.jpg",
        "item_url": "https://item.taobao.com/item.htm?id=1", "item_id": "1",
        "category": "Other",
    }
    r.update(over)
    return r


def test_build_xlsx_neutralizes_formulas_keeps_image_hyperlink(tmp_path):
    from openpyxl import load_workbook

    rows = [
        _inv_row(seller='=HYPERLINK("http://evil","shop")', title="=cmd",
                 variant="=1+1", category='=WEBSERVICE("http://evil")'),
    ]
    path = str(tmp_path / "inv.xlsx")
    build_xlsx(rows, path, embed_images=False)
    wb = load_workbook(path)
    ws = wb["Inventory"]
    # intentional IMAGE / HYPERLINK formula cells stay real formulas
    assert ws["A2"].data_type == "f" and str(ws["A2"].value).startswith("=IMAGE")
    link_col = ws.cell(1, 15).value  # header "Product link"
    assert link_col == "Product link"
    assert ws.cell(2, 15).data_type == "f" and str(ws.cell(2, 15).value).startswith("=HYPERLINK")
    # untrusted Taobao-controlled text is forced to TEXT (data_type 's'), value preserved
    for col, want in ((3, "=WEBSERVICE(\"http://evil\")"), (4, "=HYPERLINK(\"http://evil\",\"shop\")"),
                      (5, "=cmd"), (6, "=1+1")):
        cell = ws.cell(2, col)
        assert cell.data_type == "s", f"col {col} should be text, got {cell.data_type}"
        assert cell.value == want
    # no other cell may be a formula
    assert not any(c.data_type == "f" for row in ws.iter_rows(min_row=2, max_row=2)
                   for c in row if c.column not in (1, 15))
    # By Category sheet (no intentional formulas) is fully neutralized
    bycat = wb["By Category"]
    assert not any(c.data_type == "f" for row in bycat.iter_rows() for c in row)


def test_build_xlsx_embed_images_has_no_formula_cells(tmp_path):
    """embed_images=True writes thumbnails, not formulas — nothing may be executable."""
    from openpyxl import load_workbook

    rows = [_inv_row(title="=2+2", pic="")]   # no pic → no image at all
    path = str(tmp_path / "inv_embed.xlsx")
    build_xlsx(rows, path, embed_images=True)
    wb = load_workbook(path)
    ws = wb["Inventory"]
    assert not any(c.data_type == "f" for row in ws.iter_rows() for c in row)


def test_build_xlsx_sanitizes_formula_arg_breakout(tmp_path):
    """A hostile itemUrl with quotes cannot break out of the =HYPERLINK literal."""
    from openpyxl import load_workbook

    rows = [_inv_row(item_url='https://item.taobao.com/item.htm?id=1",=1+1,')]
    path = str(tmp_path / "inv_link.xlsx")
    build_xlsx(rows, path, embed_images=False)
    wb = load_workbook(path)
    link = wb["Inventory"].cell(2, 15).value
    assert isinstance(link, str) and link.startswith("=HYPERLINK(")
    assert link.count('"') == 4          # only the 4 structural quotes (url + "open") — no breakout
    inner = link[link.index('("') + 2:link.rindex('","')]
    assert inner == "https://item.taobao.com/item.htm?id=1,=1+1,"   # the whole hostile URL stays ONE string arg
    assert '",=1+1' not in link         # no quote-terminated breakout into extra args



# ── image URL policy (SSRF hardening — pure, no network) ──────────────────────
def test_image_url_policy_allows_alibaba_cdn():
    # HTTPS-only on the CDN allowlist (protocol-relative → https)
    assert _image_url_policy("//img.alicdn.com/p.jpg") == "https://img.alicdn.com/p.jpg"
    assert _image_url_policy("https://gw.alicdn.com/x.jpg") == "https://gw.alicdn.com/x.jpg"
    assert _image_url_policy("https://alicdn.com/x.jpg") == "https://alicdn.com/x.jpg"
    assert _image_url_policy("https://img.alicdn.com:443/x.jpg") == "https://img.alicdn.com:443/x.jpg"
    assert _image_url_policy("https://img.taobaocdn.com/t.jpg") == "https://img.taobaocdn.com/t.jpg"
    # cleartext http (even on the CDN allowlist) and port 80 are rejected — HTTPS-only
    assert _image_url_policy("http://alicdn.com/x.jpg") is None
    assert _image_url_policy("http://img.alicdn.com:80/x.jpg") is None


def test_image_url_policy_rejects_off_cdn_and_internal_hosts():
    for bad in (
        "http://127.0.0.1/x.jpg",                       # loopback
        "http://169.254.169.254/latest/meta-data",      # cloud metadata
        "http://localhost/x.jpg",
        "http://[::1]/x.jpg",
        "http://evil.example.com/x.jpg",
        "https://evil-alicdn.com/x.jpg",            # NOT a subdomain of alicdn.com
        "http://alicdn.com.evil.example/x.jpg",     # suffix trick: NOT *.alicdn.com
        "https://sub.alicdn.com.attacker.io/x.jpg",
        "https://alicdn.com@evil.example/x.jpg",    # userinfo + off-CDN host
    ):
        assert _image_url_policy(bad) is None, bad


def test_image_url_policy_rejects_scheme_creds_ports():
    assert _image_url_policy("file:///etc/passwd") is None
    assert _image_url_policy("ftp://img.alicdn.com/x.jpg") is None
    assert _image_url_policy("data:image/png;base64,AAAA") is None
    assert _image_url_policy("javascript:alert(1)") is None
    assert _image_url_policy("http://user:pass@img.alicdn.com/x.jpg") is None
    assert _image_url_policy("http://img.alicdn.com:8080/x.jpg") is None
    assert _image_url_policy("http://img.alicdn.com:21/x.jpg") is None
    assert _image_url_policy("") is None
    assert _image_url_policy(None) is None


def test_image_url_policy_malformed():
    assert _image_url_policy("http://img.alicdn.com:notaport/x.jpg") is None


def test_redirect_handler_blocks_off_cdn_targets():
    """A CDN URL that 302-redirects to a disallowed host must be refused (pure)."""
    import urllib.request

    req = urllib.request.Request("https://img.alicdn.com/a.jpg")
    h = _AllowedRedirectHandler()
    # redirect to an internal/metadata host → blocked
    try:
        h.redirect_request(req, None, 302, "Found", {"Location": "http://169.254.169.254/meta"},
                           "http://169.254.169.254/meta")
        assert False, "expected _UrlPolicyError"
    except _UrlPolicyError:
        pass
    # redirect to a non-CDN external host → blocked
    try:
        h.redirect_request(req, None, 302, "Found", {"Location": "http://evil.example/x"},
                           "http://evil.example/x")
        assert False, "expected _UrlPolicyError"
    except _UrlPolicyError:
        pass
    # same-CDN-host redirect → allowed (returns a follow-up Request, no exception)
    out = h.redirect_request(req, None, 302, "Found", {"Location": "b.jpg"}, "b.jpg")
    assert out is not None


def test_formula_image_url_validation_and_escaping():
    assert _formula_image_url("//img.alicdn.com/p.jpg") == "https://img.alicdn.com/p.jpg"
    assert _formula_image_url("http://evil.example/x.jpg") == ""          # off-CDN → no formula
    assert _formula_image_url("http://127.0.0.1/x.jpg") == ""
    assert _formula_image_url("") == ""
    assert _formula_image_url(None) == ""
    # untrusted quotes stripped; hostile text stays inside the URL string, not a formula
    assert _formula_image_url('//img.alicdn.com/a.jpg",=1+1,') == "https://img.alicdn.com/a.jpg,=1+1,"


def test_formula_link_url_validation_and_escaping():
    assert _formula_link_url("https://item.taobao.com/item.htm?id=1") == "https://item.taobao.com/item.htm?id=1"
    assert _formula_link_url("//detail.tmall.com/item.htm?id=2") == "https://detail.tmall.com/item.htm?id=2"
    assert _formula_link_url("http://evil.example/x") == ""               # non-Taobao → no formula
    assert _formula_link_url("http://item.taobao.com/item.htm?id=1") == ""  # HTTPS-only: http rejected
    assert _formula_link_url("https://item.taobao.com:80/item.htm?id=1") == ""  # port 80 rejected
    assert _formula_link_url("https://item.taobao.com/item.htm?id=1\",=1+1,") == \
        "https://item.taobao.com/item.htm?id=1,=1+1,"


# ── formula neutralization covers ALL injection prefixes (= + - @) ─────────────
def test_is_formula_prefix_covers_all_injection_chars():
    from src.inventory import _is_formula_prefix

    for s in ("=1+1", "+1+1", "-1", "@SUM(1,2)", "@cmd", "+cmd"):
        assert _is_formula_prefix(s) is True, s
    # leading whitespace before a formula prefix is ALSO risky (Excel/Sheets trim it) — conservative
    for s in ("  =1+1", "\t@cmd", " =1", " \n+1", "  -1", "   @cmd"):
        assert _is_formula_prefix(s) is True, repr(s)
    assert _is_formula_prefix("正常商品") is False
    assert _is_formula_prefix("") is False
    assert _is_formula_prefix(123) is False
    assert _is_formula_prefix("1+1") is False        # mid-string '+' is not an injection prefix
    assert _is_formula_prefix("价格=10元") is False   # '=' mid-string is not a formula start


def test_build_xlsx_neutralizes_plus_minus_at_prefixes(tmp_path):
    """Excel/Sheets also parse a leading + - @ as a formula — those cells must be TEXT."""
    from openpyxl import load_workbook

    rows = [_inv_row(seller="+1+1", title="-1", variant="@SUM(1,2)", category="+cmd")]
    path = str(tmp_path / "inv_prefix.xlsx")
    build_xlsx(rows, path, embed_images=True)   # no intentional formulas in embed mode
    wb = load_workbook(path)
    ws = wb["Inventory"]
    for col, want in ((3, "+cmd"), (4, "+1+1"), (5, "-1"), (6, "@SUM(1,2)")):
        cell = ws.cell(2, col)
        assert cell.data_type == "s", f"col {col} should be TEXT, got {cell.data_type}"
        assert cell.value == want
    assert not any(c.data_type == "f" for row in ws.iter_rows() for c in row)


def test_build_xlsx_neutralizes_leading_whitespace_formula(tmp_path):
    """A '  =1+1'-style cell (leading whitespace + formula prefix) is neutralized to TEXT."""
    from openpyxl import load_workbook

    rows = [_inv_row(seller="  =1+1", title="\t@cmd", variant=" +1", category="  -1")]
    path = str(tmp_path / "inv_ws.xlsx")
    build_xlsx(rows, path, embed_images=True)
    wb = load_workbook(path)
    ws = wb["Inventory"]
    for col, want in ((3, "  -1"), (4, "  =1+1"), (5, "\t@cmd"), (6, " +1")):
        cell = ws.cell(2, col)
        assert cell.data_type == "s", f"col {col} should be TEXT, got {cell.data_type}"
        assert cell.value == want
    assert not any(c.data_type == "f" for row in ws.iter_rows() for c in row)


def test_download_one_rejects_off_policy_without_network():
    """_download_one must reject off-policy URLs BEFORE any fetch or cache access — pure,
    no network (the policy returns None before urllib is ever touched)."""
    from src.inventory import _download_one

    assert _download_one("") is None
    assert _download_one("http://127.0.0.1/x.jpg") is None          # loopback
    assert _download_one("http://169.254.169.254/meta") is None     # cloud metadata
    assert _download_one("https://evil-alicdn.com/x.jpg") is None   # not a CDN subdomain
    assert _download_one("http://user:pass@img.alicdn.com/x.jpg") is None  # credentials
    assert _download_one("http://img.alicdn.com:8080/x.jpg") is None       # nonstandard port
    assert _download_one(None) is None
